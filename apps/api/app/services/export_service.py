"""Export service: CSV, Excel with conditional formatting."""

import csv
import io
from typing import Optional

from app.core.logging import get_logger

logger = get_logger(__name__)


class ExportService:
    """Generates CSV and Excel exports of grade tables."""

    def generate_csv(self, batch_name: str, sheets_data: list[dict]) -> bytes:
        """
        Generate CSV with one row per sheet.
        Columns: Roll Number, Student Name, Subject, Q1..Qn, Total, Max, Percentage, Grade, Status.
        """
        if not sheets_data:
            return b""

        # Gather all unique q_nos
        all_q_nos = set()
        for sheet in sheets_data:
            for ev in sheet.get("evaluations", []):
                all_q_nos.add(str(ev["q_no"]))
        q_nos_sorted = sorted(all_q_nos, key=lambda x: (len(x), x))

        output = io.StringIO()
        fieldnames = (
            ["Roll Number", "Student Name", "Subject"]
            + [f"Q{q}" for q in q_nos_sorted]
            + ["Total Marks", "Max Marks", "Percentage", "Grade", "Status"]
        )
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()

        for sheet in sheets_data:
            row = {
                "Roll Number": sheet.get("roll_number") or "",
                "Student Name": sheet.get("student_name") or "",
                "Subject": sheet.get("subject") or "",
                "Total Marks": sheet.get("total_awarded_marks") or 0,
                "Max Marks": sheet.get("total_max_marks") or 0,
                "Percentage": f"{sheet.get('percentage', 0):.1f}",
                "Grade": sheet.get("grade") or "",
                "Status": sheet.get("status") or "",
            }
            eval_map = {str(ev["q_no"]): ev for ev in sheet.get("evaluations", [])}
            for q in q_nos_sorted:
                ev = eval_map.get(q)
                if ev:
                    row[f"Q{q}"] = f"{ev['awarded_marks']:g}/{ev['max_marks']:g}"
                else:
                    row[f"Q{q}"] = "—"

            writer.writerow(row)

        return output.getvalue().encode("utf-8-sig")

    def generate_excel(self, batch_name: str, sheets_data: list[dict]) -> bytes:
        """
        Generate Excel with conditional formatting, bold headers, alternating rows.
        """
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side
        )
        from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
        from openpyxl.utils import get_column_letter

        if not sheets_data:
            return b""

        all_q_nos = set()
        for sheet in sheets_data:
            for ev in sheet.get("evaluations", []):
                all_q_nos.add(str(ev["q_no"]))
        q_nos_sorted = sorted(all_q_nos, key=lambda x: (len(x), x))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = batch_name[:31]

        headers = (
            ["Roll Number", "Student Name", "Subject"]
            + [f"Q{q}" for q in q_nos_sorted]
            + ["Total Marks", "Max Marks", "Percentage", "Grade", "Status"]
        )

        # Header row style
        header_fill = PatternFill("solid", fgColor="2563EB")
        header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        header_align = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # Alternating row fills
        fill_white = PatternFill("solid", fgColor="FFFFFF")
        fill_light = PatternFill("solid", fgColor="F1F5F9")

        for row_idx, sheet in enumerate(sheets_data, 2):
            row_fill = fill_white if row_idx % 2 == 0 else fill_light
            eval_map = {str(ev["q_no"]): ev for ev in sheet.get("evaluations", [])}

            row_values = (
                [
                    sheet.get("roll_number") or "",
                    sheet.get("student_name") or "",
                    sheet.get("subject") or "",
                ]
                + [
                    f"{eval_map[q]['awarded_marks']:g}/{eval_map[q]['max_marks']:g}"
                    if q in eval_map else "—"
                    for q in q_nos_sorted
                ]
                + [
                    sheet.get("total_awarded_marks") or 0,
                    sheet.get("total_max_marks") or 0,
                    float(sheet.get("percentage") or 0),
                    sheet.get("grade") or "",
                    sheet.get("status") or "",
                ]
            )

            for col_idx, value in enumerate(row_values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill

        # Conditional formatting on Percentage column
        pct_col_idx = len(headers) - 2  # 3rd from end
        pct_col_letter = get_column_letter(pct_col_idx)
        data_rows = f"{pct_col_letter}2:{pct_col_letter}{len(sheets_data) + 1}"

        green_fill = PatternFill("solid", fgColor="22C55E")
        amber_fill = PatternFill("solid", fgColor="F59E0B")
        red_fill = PatternFill("solid", fgColor="EF4444")

        ws.conditional_formatting.add(
            data_rows, CellIsRule(operator="greaterThanOrEqual", formula=["60"], fill=green_fill)
        )
        ws.conditional_formatting.add(
            data_rows,
            CellIsRule(
                operator="between", formula=["40", "59.99"], fill=amber_fill
            ),
        )
        ws.conditional_formatting.add(
            data_rows, CellIsRule(operator="lessThan", formula=["40"], fill=red_fill)
        )

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()
